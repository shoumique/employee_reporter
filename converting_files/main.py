"""
Bank Excel Bijoy→Unicode converter
===================================
Reads an Excel (.xlsx/.xls) file that may contain Bijoy-encoded Bangla text,
converts every Bijoy cell to proper Unicode, and writes:

  1. A UTF-8 CSV  : <stem>_unicode.csv
  2. A formatted Excel : <stem>_unicode.xlsx  (original styles preserved via openpyxl)

Detection strategy
------------------
unicodeconverter.convert_bijoy_to_unicode() will convert *anything* — English
words included — into Bengali codepoints.  We must decide whether to keep the
result.

Key insight (verified empirically):
  • Valid Bijoy→Unicode output always contains at least one Bengali VOWEL SIGN
    (dependent vowels, Unicode category "Mc", range 0x09BE–0x09CC).
  • Garbled English→Bengali output consists only of independent consonants and
    digits — no vowel signs — because each ASCII letter maps 1-to-1 to a
    consonant.

Therefore the rule is:
  KEEP converted result  ←→  it contains ≥ 1 Bengali vowel sign.
  DISCARD otherwise (return original string unchanged).

Additionally, strings that already contain Bengali Unicode characters are
passed through as-is (they are already correct Unicode Bangla).
"""

import argparse
import os
from typing import Any

import pandas as pd
import unicodeconverter as uc

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None


# ── Bengali Unicode ranges ─────────────────────────────────────────────────
_BN_START = 0x0980
_BN_END   = 0x09FF

# Dependent vowel signs (Mc category): 0x09BE AA, 0x09BF I, 0x09C0 II,
# 0x09C1 U, 0x09C2 UU, 0x09C3 R, 0x09C4 RR, 0x09E2 L, 0x09E3 LL,
# 0x09CB O, 0x09CC AU  — all sit in 0x09BE–0x09CC
_VOWEL_SIGN_RANGE = (0x09BE, 0x09CC)

# Khanda Ta (ৎ, U+09CE) is strictly word-final in real Bengali.
# When English text is fed through the Bijoy converter, the letter 'r'
# maps to ৎ, so English words with 'r' in a non-final position produce
# sequences like ৎৎ or ৎধ which are impossible in real Bengali.
_KHANDA_TA = '\u09CE'


def _has_bengali(text: str) -> bool:
    """True if text already contains Bengali Unicode characters."""
    return any(_BN_START <= ord(c) <= _BN_END for c in text)


def _has_vowel_sign(text: str) -> bool:
    """True if text contains at least one Bengali dependent vowel sign (Mc)."""
    return any(_VOWEL_SIGN_RANGE[0] <= ord(c) <= _VOWEL_SIGN_RANGE[1] for c in text)


def _has_invalid_khanda_ta(text: str) -> bool:
    """True if ৎ is immediately followed by another Bengali codepoint.

    In valid Bengali, ৎ (Khanda Ta) only ever appears at the end of a word —
    never before another Bengali letter, sign, or digit. Its presence before
    another Bengali codepoint is a reliable indicator that the text is garbled
    output produced by running English through the Bijoy converter (the letter
    'r' in Bijoy maps to ৎ).
    """
    for i, ch in enumerate(text):
        if ch == _KHANDA_TA and i + 1 < len(text):
            if _BN_START <= ord(text[i + 1]) <= _BN_END:
                return True
    return False


def convert_bijoy_in_value(value: Any) -> Any:
    """Convert a single cell value from Bijoy to Unicode when appropriate.

    Rules applied in order:
    1. Non-strings pass through unchanged.
    2. Strings already containing Bengali Unicode pass through unchanged.
    3. Try unicodeconverter on the string.
    4. Keep the converted result ONLY if it contains a Bengali vowel sign —
       this reliably separates valid Bangla from garbled English.
    5. Otherwise return the original value unchanged.
    """
    if not isinstance(value, str):
        return value
    if not value.strip():
        return value

    # Already proper Unicode Bangla — leave alone
    if _has_bengali(value):
        return value

    try:
        converted = uc.convert_bijoy_to_unicode(value)
    except Exception:
        return value

    # Accept only if conversion produced valid Bangla:
    #   1. Has at least one Bengali dependent vowel sign (rules out pure-consonant garble).
    #   2. ৎ does NOT appear before another Bengali codepoint (rules out English
    #      words containing 'r', which Bijoy maps to ৎ, e.g. "Super-Newmerray").
    if _has_vowel_sign(converted) and not _has_invalid_khanda_ta(converted):
        return converted

    # Otherwise: garbled output from English or non-Bijoy text — discard.
    return value


# ── Output helpers ─────────────────────────────────────────────────────────

def _apply_to_df(df: pd.DataFrame) -> pd.DataFrame:
    """Apply Bijoy conversion to all cells, column headers, and index."""
    df_out = df.map(convert_bijoy_in_value)
    df_out.columns = [convert_bijoy_in_value(c) for c in df_out.columns]
    df_out.index   = [convert_bijoy_in_value(i) for i in df_out.index]
    return df_out


def write_csv(df: pd.DataFrame, path: str) -> None:
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    df.to_csv(path, index=False, encoding="utf-8-sig")


def write_excel_from_df(df: pd.DataFrame, path: str) -> None:
    """Write a DataFrame to .xlsx using openpyxl (no original formatting)."""
    if load_workbook is None:
        raise RuntimeError("Install openpyxl:  pip install openpyxl")
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sheet1")


def write_excel_preserve_formatting(input_xlsx: str, output_xlsx: str) -> None:
    """Load original Excel with openpyxl, convert cell text, save as new file.

    All cell styles, column widths, merged cells, etc. are preserved because
    we only replace .value on string cells.
    """
    if load_workbook is None:
        raise RuntimeError("Install openpyxl:  pip install openpyxl")

    wb = load_workbook(input_xlsx, data_only=False)

    for ws in wb.worksheets:
        ws.title = convert_bijoy_in_value(ws.title)
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and not cell.value.startswith("="):
                    cell.value = convert_bijoy_in_value(cell.value)

    os.makedirs(os.path.dirname(output_xlsx) or ".", exist_ok=True)
    wb.save(output_xlsx)


# ── CLI ────────────────────────────────────────────────────────────────────

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Convert Bijoy-encoded Bangla text in an Excel file to Unicode. "
            "Produces a UTF-8 CSV and a formatted Excel copy."
        ),
    )
    parser.add_argument("input_file", help="Input Excel (.xlsx/.xls) or CSV file.")
    parser.add_argument(
        "-o", "--output-dir",
        default=None,
        help="Directory to write outputs (default: same directory as input).",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_path = os.path.abspath(args.input_file)

    if not os.path.isfile(input_path):
        raise FileNotFoundError(f"Input file not found: {input_path}")

    stem, ext = os.path.splitext(input_path)
    out_dir   = os.path.abspath(args.output_dir) if args.output_dir else os.path.dirname(input_path)
    base_name = os.path.basename(stem)

    out_csv   = os.path.join(out_dir, f"{base_name}_unicode.csv")
    out_xlsx  = os.path.join(out_dir, f"{base_name}_unicode.xlsx")

    if ext.lower() in (".xlsx", ".xls"):
        # ── Excel input path ──────────────────────────────────────────────
        # 1. Write CSV (via pandas — no formatting)
        df = pd.read_excel(input_path, dtype=str, keep_default_na=False)
        df_converted = _apply_to_df(df)
        write_csv(df_converted, out_csv)
        print(f"✓ CSV written to:   {out_csv}")

        # 2. Write formatting-preserved Excel (via openpyxl direct edit)
        if load_workbook is not None:
            write_excel_preserve_formatting(input_path, out_xlsx)
            print(f"✓ Excel written to: {out_xlsx}")
        else:
            # Fallback — no formatting preserved but still useful
            write_excel_from_df(df_converted, out_xlsx)
            print(f"✓ Excel written to: {out_xlsx}  (openpyxl not found; formatting not preserved)")

    elif ext.lower() == ".csv":
        # ── CSV input path ────────────────────────────────────────────────
        df = pd.read_csv(input_path, dtype=str, keep_default_na=False)
        df_converted = _apply_to_df(df)
        write_csv(df_converted, out_csv)
        print(f"✓ CSV written to:   {out_csv}")

        write_excel_from_df(df_converted, out_xlsx)
        print(f"✓ Excel written to: {out_xlsx}")

    else:
        raise ValueError(f"Unsupported file type: {ext}. Use .xlsx, .xls, or .csv")


if __name__ == "__main__":
    main()

